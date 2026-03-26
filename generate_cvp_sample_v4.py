"""
╔══════════════════════════════════════════════════════════════════╗
║  CVP_Sample_Data_V4 — 範例資料生成腳本                          ║
║  輸出：CVP_Sample_Data_V4.xlsx（10 個工作表）                   ║
║  ─────────────────────────────────────────────────────────────  ║
║  V3 相容（6張）：Product_Mix / Parameters / Investment /        ║
║                  Capital_Efficiency / Monthly_Trend /           ║
║                  Risk_Parameters                                 ║
║  V4 新增（4張）：Budget_Data / CashFlow_Params /                ║
║                  YoY_Comparison / Scenario_Weights              ║
╚══════════════════════════════════════════════════════════════════╝
"""

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import random
import math

# ─────────────────────────────────────────────
# 顏色常數
# ─────────────────────────────────────────────
C_HDR_BLUE   = "1E3A5F"   # 深藍（V3/V4 共用表頭）
C_HDR_NEW    = "1E4D3B"   # 深綠（V4 新增表頭）
C_ACCENT     = "0EA5E9"   # 天藍
C_PURPLE     = "7C3AED"   # 紫
C_EMERALD    = "059669"   # 綠
C_AMBER      = "D97706"   # 琥珀
C_RED        = "DC2626"   # 紅
C_ROW_ODD    = "F8FAFC"   # 淡藍白交替列
C_ROW_EVEN   = "FFFFFF"   # 白
C_TOTAL      = "EFF6FF"   # 合計列底色
C_NOTE       = "FEF3C7"   # 說明列底色

def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font(color="FFFFFF", bold=True, size=11):
    return Font(name="微軟正黑體", color=color, bold=bold, size=size)

def _hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _fmt_sheet(ws, col_widths, freeze="A2"):
    """套用通用格式：凍結首列、欄寬"""
    ws.freeze_panes = freeze
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _write_header(ws, row, headers, bg_color=C_HDR_BLUE):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font      = _hdr_font()
        cell.fill      = _hdr_fill(bg_color)
        cell.alignment = _center()
        cell.border    = _thin_border()

def _write_row(ws, row, values, bg_color=None, bold=False, num_fmt=None):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font      = Font(name="微軟正黑體", bold=bold, size=10)
        cell.alignment = _center()
        cell.border    = _thin_border()
        if bg_color:
            cell.fill = _hdr_fill(bg_color)
    return ws

def _note_row(ws, row, text, ncols, bg_color=C_NOTE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(name="微軟正黑體", size=9, color="92400E", italic=True)
    cell.fill      = _hdr_fill(bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _thin_border()
    ws.row_dimensions[row].height = 22

# ─────────────────────────────────────────────
# 基礎業務資料（與 V3 相同產品，但欄位擴充）
# ─────────────────────────────────────────────
PRODUCTS = [
    # name,               price,   vc,      vol,  capMax, minVol, cat
    ("旗艦型伺服器 (A)",  500000,  280000,  50,   80,     10,     "硬體產品"),
    ("高階工作站 (B)",    150000,  95000,   200,  300,    20,     "硬體產品"),
    ("商務筆電 (C)",      45000,   32000,   1000, 1500,   100,    "硬體產品"),
    ("入門型筆電 (D)",    20000,   15000,   2500, 4000,   500,    "硬體產品"),
    ("軟體授權 (E)",      8000,    1200,    3000, 9999,   0,      "軟體服務"),
    ("維護合約 (F)",      15000,   3000,    800,  9999,   0,      "軟體服務"),
]

MONTHLY_SEASON = [0.80, 0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15, 1.08, 1.12, 1.20, 1.25]
MONTHS = ["1月","2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月"]
FC_ANNUAL    = 25_000_000   # 固定成本（元）
FC_MONTHLY   = FC_ANNUAL // 12

wb = openpyxl.Workbook()
wb.remove(wb.active)  # 移除預設工作表

# ══════════════════════════════════════════════
# 1. Product_Mix（V3 相容 + 新增最低銷量欄）
# ══════════════════════════════════════════════
ws1 = wb.create_sheet("Product_Mix")
ws1.sheet_properties.tabColor = "0EA5E9"

headers1 = ["產品名稱", "單價（元）", "單位變動成本（元）", "預估銷量（件）",
            "產能上限（件）", "最低銷量（件）", "產品類別"]
ws1.row_dimensions[1].height = 28
_write_header(ws1, 1, headers1)

for i, (name, price, vc, vol, cap, minvol, cat) in enumerate(PRODUCTS, 2):
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws1, i, [name, price, vc, vol, cap, minvol, cat], bg_color=bg)
    # 左對齊名稱
    ws1.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    # 數值格式
    for col in [2, 3]:
        ws1.cell(i, col).number_format = '#,##0'
    for col in [4, 5, 6]:
        ws1.cell(i, col).number_format = '#,##0'

# 說明行
_note_row(ws1, len(PRODUCTS)+2, " ※ 單價 / 單位變動成本單位為「元」；銷量單位為「件」", 7)
_fmt_sheet(ws1, [22, 14, 18, 14, 14, 14, 14])

# ══════════════════════════════════════════════
# 2. Parameters（V3 + V4 新增參數）
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Parameters")
ws2.sheet_properties.tabColor = "8B5CF6"

headers2 = ["參數名稱", "數值", "單位", "說明"]
ws2.row_dimensions[1].height = 28
_write_header(ws2, 1, headers2)

params2 = [
    ("固定成本", FC_ANNUAL, "元/年", "含租金、人事、折舊等固定支出"),
    ("目標利潤", 15_000_000, "元/年", "年度利潤目標"),
    ("安全邊際警示門檻", 0.15, "比例（0-1）", "低於此值顯示警示"),
    ("樂觀情境銷量增幅", 0.20, "比例", "樂觀情境下銷量增幅假設"),
    ("悲觀情境銷量減幅", -0.15, "比例", "悲觀情境下銷量減幅假設（負值）"),
    ("競爭衝擊價格降幅", -0.10, "比例", "受競爭衝擊時的價格下降幅度（負值）"),
    ("邊際貢獻率優良門檻", 0.40, "比例", "達到此 CM 率視為優良"),
    ("行業平均邊際貢獻率", 0.35, "比例", "所在行業的平均 CM 率基準"),
    ("公司名稱", "範例科技股份有限公司", "文字", "顯示於報告頁首"),
    ("期間", "2026 年度", "文字", "分析期間，顯示於頁首"),
]

for i, (nm, val, unit, desc) in enumerate(params2, 2):
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws2, i, [nm, val, unit, desc], bg_color=bg)
    ws2.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws2.cell(i, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if isinstance(val, float):
        ws2.cell(i, 2).number_format = '0.00'

_note_row(ws2, len(params2)+2, " ※ 比例類參數使用小數（如 0.15 = 15%）；金額類使用「元」為單位", 4)
_fmt_sheet(ws2, [28, 20, 16, 38])

# ══════════════════════════════════════════════
# 3. Investment（V3 相容）
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("Investment")
ws3.sheet_properties.tabColor = "F59E0B"

headers3 = ["參數名稱", "數值", "單位", "說明"]
ws3.row_dimensions[1].height = 28
_write_header(ws3, 1, headers3)

params3 = [
    ("六年研發暨生產線投資（Capex）", 120_000_000, "元", "五年資本支出總額"),
    ("投資年期",                      5,            "年", "計算 NPV/IRR 的期間"),
    ("WACC（加權平均資金成本）",       0.08,         "比例（0-1）", "折現率，如 0.08 = 8%"),
    ("折舊年限（depLife）",            5,            "年", "直線折舊年限"),
    ("殘值（salvage）比例",            0.10,         "比例（0-1）", "資產殘值佔原值比例"),
]

for i, (nm, val, unit, desc) in enumerate(params3, 2):
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws3, i, [nm, val, unit, desc], bg_color=bg)
    ws3.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws3.cell(i, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if isinstance(val, int) and val > 1000:
        ws3.cell(i, 2).number_format = '#,##0'
    elif isinstance(val, float):
        ws3.cell(i, 2).number_format = '0.00'

_note_row(ws3, len(params3)+2, " ※ 投資金額單位為「元」；比例類參數使用小數（WACC=0.08 表示 8%）", 4)
_fmt_sheet(ws3, [32, 18, 16, 38])

# ══════════════════════════════════════════════
# 4. Capital_Efficiency（V3 相容）
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("Capital_Efficiency")
ws4.sheet_properties.tabColor = "10B981"

headers4 = ["參數名稱", "數值", "單位", "說明"]
ws4.row_dimensions[1].height = 28
_write_header(ws4, 1, headers4)

params4 = [
    ("流動資產",         150_000_000, "元", "含現金、應收帳款、存貨等"),
    ("固定資產",         280_000_000, "元", "廠房、設備等長期資產"),
    ("其他資產",         70_000_000,  "元", "無形資產、長期投資等"),
    ("流動負債",         80_000_000,  "元", "一年內到期負債"),
    ("長期負債",         120_000_000, "元", "一年以上到期負債"),
    ("股東權益",         300_000_000, "元", "資本額＋保留盈餘"),
    ("行業平均資產周轉率", 1.20,      "次/年", "所在行業標準值"),
    ("行業平均 ROE",     0.15,        "比例（0-1）", "行業平均股東權益報酬率"),
    ("行業平均 ROA",     0.08,        "比例（0-1）", "行業平均資產報酬率"),
]

for i, (nm, val, unit, desc) in enumerate(params4, 2):
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws4, i, [nm, val, unit, desc], bg_color=bg)
    ws4.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws4.cell(i, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if isinstance(val, int) and val > 1000:
        ws4.cell(i, 2).number_format = '#,##0'
    elif isinstance(val, float):
        ws4.cell(i, 2).number_format = '0.000'

_note_row(ws4, len(params4)+2, " ※ 所有金額單位為「元」；比例類參數使用小數（0.15 = 15%）", 4)
_fmt_sheet(ws4, [24, 18, 16, 38])

# ══════════════════════════════════════════════
# 5. Monthly_Trend（V3 相容，12 個月真實計算）
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("Monthly_Trend")
ws5.sheet_properties.tabColor = "06B6D4"

headers5 = ["月份", "總收入（元）", "總變動成本（元）", "固定成本（元）",
            "邊際貢獻（元）", "營業利潤（元）", "邊際貢獻率"]
ws5.row_dimensions[1].height = 28
_write_header(ws5, 1, headers5)

# 計算基準收入與 VC
base_rev = sum(p[1] * p[3] for p in PRODUCTS)   # 年度基準總收入
base_vc  = sum(p[2] * p[3] for p in PRODUCTS)   # 年度基準總變動成本

random.seed(42)
annual_rev = annual_cm = annual_profit = 0

for m_idx, (month, season) in enumerate(zip(MONTHS, MONTHLY_SEASON), 2):
    noise = 1 + random.uniform(-0.03, 0.03)
    rev   = round(base_rev  / 12 * season * noise)
    vc    = round(base_vc   / 12 * season * noise * (1 + random.uniform(-0.02, 0.02)))
    cm    = rev - vc
    profit= cm - FC_MONTHLY
    cmrate= round(cm / rev, 4) if rev else 0
    annual_rev    += rev
    annual_cm     += cm
    annual_profit += profit
    bg = C_ROW_ODD if m_idx % 2 == 0 else C_ROW_EVEN
    _write_row(ws5, m_idx, [month, rev, vc, FC_MONTHLY, cm, profit, cmrate], bg_color=bg)
    for col in [2, 3, 4, 5, 6]:
        ws5.cell(m_idx, col).number_format = '#,##0'
    ws5.cell(m_idx, 7).number_format = '0.00%'

# 年度合計列
r_total = 14
_write_row(ws5, r_total,
    ["年度合計", annual_rev, annual_rev - annual_cm, FC_ANNUAL,
     annual_cm, annual_profit, round(annual_cm/annual_rev, 4) if annual_rev else 0],
    bg_color=C_TOTAL, bold=True)
for col in [2, 3, 4, 5, 6]:
    ws5.cell(r_total, col).number_format = '#,##0'
ws5.cell(r_total, 7).number_format = '0.00%'

_note_row(ws5, 15, " ※ 月度收入依季節指數（Q1低/Q4高）× 隨機擾動生成；金額單位為「元」", 7)
_fmt_sheet(ws5, [8, 16, 16, 14, 16, 16, 12])

# ══════════════════════════════════════════════
# 6. Risk_Parameters（V3 相容）
# ══════════════════════════════════════════════
ws6 = wb.create_sheet("Risk_Parameters")
ws6.sheet_properties.tabColor = "EF4444"

headers6 = ["風險事件", "發生機率（0-1）", "財務衝擊度（0-1）", "分類", "應對措施"]
ws6.row_dimensions[1].height = 28
_write_header(ws6, 1, headers6)

risks = [
    ("主要原材料漲價 15%",    0.45, 0.65, "成本風險",   "多供應商備份策略，提前鎖定長約"),
    ("核心客戶流失",          0.25, 0.80, "收入風險",   "強化客戶關係，開發替代客戶"),
    ("新競爭者市場進入",      0.60, 0.50, "競爭風險",   "加速差異化，提升轉換成本"),
    ("匯率大幅波動（±10%）", 0.40, 0.40, "財務風險",   "外匯避險對沖策略"),
    ("關鍵人才離職",          0.30, 0.35, "營運風險",   "人才保留計畫及接班培育"),
    ("法規政策重大改變",      0.20, 0.55, "法規風險",   "法規監控機制，調整產品合規"),
    ("資訊系統崩潰",          0.15, 0.45, "技術風險",   "DR 備援系統，定期演練"),
    ("銀行信用額度收縮",      0.20, 0.60, "流動性風險", "備用信貸額度，現金流預先管理"),
    ("供應鏈斷鏈 > 2 週",   0.25, 0.70, "供應風險",   "備料安全庫存 4 週，多元物流"),
    ("產品品質重大瑕疵",      0.10, 0.85, "聲譽風險",   "品控稽核，產品責任險"),
]

for i, (nm, prob, impact, cat, mit) in enumerate(risks, 2):
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws6, i, [nm, prob, impact, cat, mit], bg_color=bg)
    ws6.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws6.cell(i, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws6.cell(i, 2).number_format = '0.00'
    ws6.cell(i, 3).number_format = '0.00'
    # 高風險標紅
    score = prob * impact
    if score > 0.30:
        ws6.cell(i, 1).font = Font(name="微軟正黑體", size=10, color=C_RED)

_note_row(ws6, len(risks)+2, " ※ 機率與衝擊度均為 0~1 之間；風險分數 = 機率 × 衝擊度 > 0.3 標示為高風險", 5)
_fmt_sheet(ws6, [26, 16, 16, 14, 36])

# ══════════════════════════════════════════════
# 7. Budget_Data（V4 新增）
# ══════════════════════════════════════════════
ws7 = wb.create_sheet("Budget_Data")
ws7.sheet_properties.tabColor = "7C3AED"

headers7 = ["產品名稱", "預算單價（元）", "預算單位變動成本（元）", "預算銷量（件）", "產品類別", "說明"]
ws7.row_dimensions[1].height = 28
_write_header(ws7, 1, headers7, bg_color=C_HDR_NEW)

for i, (name, price, vc, vol, cap, minvol, cat) in enumerate(PRODUCTS, 2):
    # 預算略高於實際（模擬預算與實際的差異）
    bprice = round(price * 1.02)
    bvc    = round(vc    * 0.97)
    bvol   = round(vol   * 0.95)
    diff   = "量略低於預算" if vol < bvol else "超出預算 ✓"
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws7, i, [name, bprice, bvc, bvol, cat, diff], bg_color=bg)
    ws7.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws7.cell(i, 6).alignment = Alignment(horizontal="left", vertical="center")
    ws7.cell(i, 2).number_format = '#,##0'
    ws7.cell(i, 3).number_format = '#,##0'

_note_row(ws7, len(PRODUCTS)+2,
    " ※ V4 新增工作表：提供預算目標值，系統將與 Product_Mix 中的實際值進行差異比較（T14 預算差異分析）", 6, "D1FAE5")
_fmt_sheet(ws7, [22, 18, 20, 14, 14, 18])

# ══════════════════════════════════════════════
# 8. CashFlow_Params（V4 新增）
# ══════════════════════════════════════════════
ws8 = wb.create_sheet("CashFlow_Params")
ws8.sheet_properties.tabColor = "0891B2"

headers8 = ["參數名稱", "數值", "單位", "說明"]
ws8.row_dimensions[1].height = 28
_write_header(ws8, 1, headers8, bg_color=C_HDR_NEW)

params8 = [
    ("應收帳款天數（AR Days）",   45,          "天",    "平均收款週期，影響現金流入時點"),
    ("應付帳款天數（AP Days）",   30,          "天",    "平均付款週期，影響現金流出時點"),
    ("最低現金安全水位",          5_000_000,   "元",    "低於此值系統顯示危急警示"),
    ("期初現金部位",              10_000_000,  "元",    "試算期初可用現金"),
]

for i, (nm, val, unit, desc) in enumerate(params8, 2):
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws8, i, [nm, val, unit, desc], bg_color=bg)
    ws8.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws8.cell(i, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if isinstance(val, int) and val > 100:
        ws8.cell(i, 2).number_format = '#,##0'

_note_row(ws8, len(params8)+2,
    " ※ V4 新增工作表：現金流預測參數（T15 現金流量預測模組使用）", 4, "D1FAE5")
_fmt_sheet(ws8, [30, 18, 12, 40])

# ══════════════════════════════════════════════
# 9. YoY_Comparison（V4 新增）── 近 3 年月度
# ══════════════════════════════════════════════
ws9 = wb.create_sheet("YoY_Comparison")
ws9.sheet_properties.tabColor = "059669"

headers9 = ["年度", "月份", "總收入（元）", "邊際貢獻（元）", "營業利潤（元）"]
ws9.row_dimensions[1].height = 28
_write_header(ws9, 1, headers9, bg_color=C_HDR_NEW)

YEARS = [("2024", 0.82), ("2025", 0.92), ("2026", 1.00)]
yr_colors = {"2024": "EDE9FE", "2025": "E0F2FE", "2026": C_ROW_ODD}
random.seed(2024)

row_r9 = 2
for year_str, yr_factor in YEARS:
    for m_idx, (month, season) in enumerate(zip(MONTHS, MONTHLY_SEASON)):
        noise  = 1 + random.uniform(-0.04, 0.04)
        rev    = round(base_rev / 12 * season * yr_factor * noise)
        vc_m   = round(base_vc  / 12 * season * yr_factor * noise * (1 + random.uniform(-0.02, 0.02)))
        cm     = rev - vc_m
        profit = cm - FC_MONTHLY
        bg = yr_colors.get(year_str, C_ROW_ODD)
        _write_row(ws9, row_r9, [year_str, month, rev, cm, profit], bg_color=bg)
        ws9.cell(row_r9, 1).alignment = Alignment(horizontal="center", vertical="center")
        for col in [3, 4, 5]:
            ws9.cell(row_r9, col).number_format = '#,##0'
        if profit < 0:
            ws9.cell(row_r9, 5).font = Font(name="微軟正黑體", size=10, color=C_RED)
        row_r9 += 1

_note_row(ws9, row_r9,
    " ※ V4 新增工作表：多年度月度資料，支援 T18 多年度比較（YoY）及 CAGR 計算", 5, "D1FAE5")
_fmt_sheet(ws9, [10, 8, 16, 16, 16], freeze="C2")

# ══════════════════════════════════════════════
# 10. Scenario_Weights（V4 新增）
# ══════════════════════════════════════════════
ws10 = wb.create_sheet("Scenario_Weights")
ws10.sheet_properties.tabColor = "EC4899"

headers10 = ["情境名稱", "銷量變動%", "單價變動%", "變動成本變動%", "固定成本變動%",
             "發生機率（0-1）", "說明"]
ws10.row_dimensions[1].height = 28
_write_header(ws10, 1, headers10, bg_color=C_HDR_NEW)

scenarios10 = [
    ("基準情境",   0,    0,    0,    0,    0.50, "當前趨勢維持，各項參數不變"),
    ("樂觀情境",   20,   0,    0,    0,    0.20, "市場需求強勁，銷量增加 20%"),
    ("悲觀情境",  -15,   0,    5,    0,    0.15, "需求萎縮，成本因通膨上升 5%"),
    ("競爭衝擊",   0,   -10,   0,    0,    0.10, "競爭者削價，單價被迫下降 10%"),
    ("成本激升",   0,    0,   15,    5,    0.05, "原物料大漲 15%，固定成本增 5%"),
]

for i, (nm, vol, price, vc, fc, prob, desc) in enumerate(scenarios10, 2):
    bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
    _write_row(ws10, i, [nm, vol, price, vc, fc, prob, desc], bg_color=bg)
    ws10.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws10.cell(i, 7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws10.cell(i, 6).number_format = '0.00'
    for col in [2, 3, 4, 5]:
        ws10.cell(i, col).number_format = '+0;-0;0'

# 機率合計驗證列
total_prob = sum(s[5] for s in scenarios10)
r_tot10 = len(scenarios10) + 2
_write_row(ws10, r_tot10, ["機率合計", "", "", "", "", total_prob, "應 = 1.00"],
    bg_color=C_TOTAL, bold=True)
ws10.cell(r_tot10, 6).number_format = '0.00'
prob_ok = abs(total_prob - 1.0) < 0.001
ws10.cell(r_tot10, 7).value = "✅ 合計符合" if prob_ok else "⚠ 請調整至合計 1.00"
ws10.cell(r_tot10, 7).font = Font(name="微軟正黑體", size=10,
    color=C_EMERALD if prob_ok else C_RED, bold=True)

_note_row(ws10, r_tot10+1,
    " ※ V4 新增工作表：各情境機率之和需等於 1.00；系統計算加權期望利潤（T5 情境分析）", 7, "D1FAE5")
_fmt_sheet(ws10, [18, 12, 12, 15, 15, 16, 36])

# ══════════════════════════════════════════════
# 封面說明工作表
# ══════════════════════════════════════════════
ws0 = wb.create_sheet("📋 使用說明", 0)
ws0.sheet_properties.tabColor = "1E3A5F"
ws0.column_dimensions["A"].width = 6
ws0.column_dimensions["B"].width = 28
ws0.column_dimensions["C"].width = 55
ws0.column_dimensions["D"].width = 20

# 標題
ws0.merge_cells("A1:D1")
t = ws0["A1"]
t.value = "CVP V4.0 — 範例資料說明書"
t.font  = Font(name="微軟正黑體", size=18, bold=True, color="FFFFFF")
t.fill  = PatternFill("solid", fgColor=C_HDR_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[1].height = 38

ws0.merge_cells("A2:D2")
t2 = ws0["A2"]
t2.value = "共 10 個工作表：6 張與 V3 格式相容，4 張 V4 新增。請依欄位說明填入真實數據後匯入儀表板。"
t2.font  = Font(name="微軟正黑體", size=11, color="92400E")
t2.fill  = PatternFill("solid", fgColor=C_NOTE)
t2.alignment = Alignment(horizontal="left", vertical="center")
ws0.row_dimensions[2].height = 26

headers0 = ["#", "工作表名稱", "用途說明", "版本"]
for ci, h in enumerate(headers0, 1):
    c = ws0.cell(4, ci, h)
    c.font = _hdr_font()
    c.fill = _hdr_fill(C_HDR_BLUE)
    c.alignment = _center()
    c.border = _thin_border()
ws0.row_dimensions[4].height = 26

sheets_info = [
    (1, "Product_Mix",        "產品售價、變動成本、預估銷量、產能上限（V3 + 新增最低銷量欄）", "V3+"),
    (2, "Parameters",         "系統全域參數：固定成本、目標利潤、警示門檻、公司資訊等",        "V3+"),
    (3, "Investment",         "投資計畫參數：Capex、WACC、折舊年限、殘值比例等",              "V3"),
    (4, "Capital_Efficiency", "資本效率參數：資產負債表科目、行業均值基準等",                  "V3"),
    (5, "Monthly_Trend",      "月度損益資料（12 個月）：收入、CM、固定成本、利潤",             "V3"),
    (6, "Risk_Parameters",    "風險事件清單：機率、衝擊度、分類、應對措施",                    "V3"),
    (7, "Budget_Data",        "【V4 新增】預算對照：各產品的預算單價/成本/銷量",              "V4 ★"),
    (8, "CashFlow_Params",    "【V4 新增】現金流預測參數：AR天數、AP天數、安全水位",           "V4 ★"),
    (9, "YoY_Comparison",     "【V4 新增】多年度月度資料（近 3 年），用於 YoY 比較",           "V4 ★"),
    (10,"Scenario_Weights",   "【V4 新增】情境機率加權設定（需合計 = 1.0）",                  "V4 ★"),
]

for i, (no, name, desc, ver) in enumerate(sheets_info, 5):
    is_v4 = "V4" in ver
    bg = "D1FAE5" if is_v4 else (C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN)
    for ci, val in enumerate([no, name, desc, ver], 1):
        c = ws0.cell(i, ci, val)
        c.font   = Font(name="微軟正黑體", size=10,
                        bold=is_v4,
                        color=C_EMERALD if is_v4 else "1E293B")
        c.fill   = _hdr_fill(bg)
        c.alignment = Alignment(horizontal="left" if ci >= 2 else "center",
                                vertical="center", wrap_text=True)
        c.border = _thin_border()
    ws0.row_dimensions[i].height = 30 if is_v4 else 24

# ─────────────────────────────────────────────
# 儲存
# ─────────────────────────────────────────────
OUT = r"c:\Users\jamic\本量利分析\CVP_Sample_Data_V4.xlsx"
wb.save(OUT)
print(f"✅ 已生成：{OUT}")
print(f"   工作表：{wb.sheetnames}")
